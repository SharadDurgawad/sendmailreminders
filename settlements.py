#-------------------------------------------------------------------------------
# Name:         settlements.py
# Purpose:      This program checks the payments not done by the members and send
#               the mails to those members
#
# Author:       sdurgawad
#
# Created:      23/07/2016
# Copyright:    (c) sdurgawad 2016
# Licence:      <your licence>
#-------------------------------------------------------------------------------

import openpyxl, smtplib, sys
from datetime import datetime

# This module will be used for sending the SMS to the members
from twilio.rest import TwilioRestClient

def sendSMStoMembers(unpaidsmsMembers, latestMonth):
    """ This function sends the SMS to the unpaid members """

    accountSID = 'AC8cbd4a333c67c10d8ae9b7f4d91f0916'
    authToken = '2cb23c3f21ff403b8168dde427dd540b'

    twilioCli = TwilioRestClient(accountSID, authToken)

    myTwilioNumber = '+19183763736' # This is the Twilio number


    for name, mobile in unpaidsmsMembers.items():
        body = "Dear %s,\n\nRecords show that you have not paid dues for %s. \
                \n\nPlease make this payment as soon as possible. \
                \n\nThank you! \n\nRegards, \n\nSharad Durgawad" % (name, latestMonth)

        print('Sending sms to %s...' % mobile)

        # Add +91 to the beginning of mobile number

        mobile = '+91' + str(mobile)

        # Send the sms to the mobiles
        message = twilioCli.messages.create(to=mobile, from_=myTwilioNumber, body=body)




def sendMailtoMembers(unpaidMembers, latestMonth):
    """ This function sends the mail to the unpaid members """

    mailFrom = 'durgawad@gmail.com'

    smtpObj = smtplib.SMTP('smtp.gmail.com', 587)

    smtpObj.ehlo()
    smtpObj.starttls()

    password = raw_input("\n Enter the password")

    smtpObj.login(mailFrom, password)

    for name, email in unpaidMembers.items():
        body = "Subject: %s dues unpaid.\nDear %s,\n\nRecords show that you have not paid dues for %s. \
                \n\nPlease make this payment as soon as possible. \
                \n\nThank you! \n\nRegards, \n\nSharad" % (latestMonth, name, latestMonth)

        print('Sending email to %s...' % email)

        # Send the mail to email
        sendmailStatus = smtpObj.sendmail(mailFrom, email, body)

        if sendmailStatus != {}:
           print('There was a problem sending email to %s: %s' % (email, sendmailStatus))


    smtpObj.quit()



def convertDate(date):
    """ This function converts the date into the format mmm YYYY \
        for example, Jun 2016 """

    #split the strings
    date=date.split('/')

    #day
    day=date[0]

    #create a dictionary for the months
    monthDict={1:'Jan', 2:'Feb', 3:'Mar', 4:'Apr', 5:'May', 6:'Jun', 7:'Jul', 8:'Aug', 9:'Sep', 10:'Oct', 11:'Nov', 12:'Dec'}

    #month
    monthIndex= int(date[1])

    month = monthDict[monthIndex]
    #year
    year=date[2]

    return month, year


def main():
    """ This is the main method where the program begins """

    wb = openpyxl.load_workbook('settlements.xlsx')

    sheet = wb.get_sheet_by_name('Sheet1')

    # Check the payment status of each member

    latestmonth = sheet.cell(row = 1, column = sheet.max_column).value

    # convertDate(str(latestmonth))

    d = datetime.strptime(str(latestmonth), '%Y-%m-%d %H:%M:%S')
    day_string = d.strftime('%d/%m/%Y')

    # Call the function convertDate to get the values in mmm YYYY format
    month, year = convertDate(str(day_string))

    # Declare the dictionary for the list of unpaid members
    unpaidMembers = {}
    unpaidsmsMembers = {}

    for i in range(2, sheet.max_row + 1):

        # get the payment status from the last column
        paymentStatus = sheet.cell(row = i, column = sheet.max_column - 1).value

        # If not paid then send the mail to the member
        if paymentStatus <> 'paid':
            name = sheet.cell(row = i, column = 1).value
            email = sheet.cell(row = i, column = 2).value
            amountDue = sheet.cell(row = i, column = 4).value
            mobile = sheet.cell(row = i, column = 3).value
            unpaidMembers[name] = email
            unpaidsmsMembers[name] = mobile

    # call the sendMailtoMembers function to send the reminder mails to unpaid members
    # month + ' ' + year is in the form of Jan 2016
    sendMailtoMembers(unpaidMembers, month + ' ' + year)

    # call the sendSMStoMembers function to send the reminder sms to unpaid members
    # month + ' ' + year is in the form of Jan 2016
    sendSMStoMembers(unpaidsmsMembers, month + ' ' + year)

    sys.exit()

main()
